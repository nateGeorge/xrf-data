from openpyxl import load_workbook
import csv, re, os, glob
import numpy as np

def interp_to_eff(eff_data_DW,dataset_DW,dataset):
	min_effdw=min(eff_data_DW)
	max_effdw=max(eff_data_DW)

	min_eff_dw_index=min(range(len(dataset_DW)), key=lambda i: 
	abs(dataset_DW[i]-min_effdw))
	max_eff_dw_index=min(range(len(dataset_DW)), key=lambda i: 
	abs(dataset_DW[i]-max_effdw))

	dataset_dw_index_min=min(range(len(dataset_DW)), key=lambda i: 
	abs(dataset_DW[i]-min_effdw))
	dataset_dw_index_max=min(range(len(dataset_DW)), key=lambda i: 
	abs(dataset_DW[i]-max_effdw))
	
	if (dataset_dw_index_max-dataset_dw_index_min)>(max_eff_dw_index-min_eff_dw_index):
		dataset_dw_index_max-=1
	if (dataset_dw_index_max-dataset_dw_index_min)<(max_eff_dw_index-min_eff_dw_index):
		dataset_dw_index_max+=1
	
	interped_data=np.interp(eff_data_DW,
	dataset_DW[min_eff_dw_index:max_eff_dw_index],
	dataset[dataset_dw_index_min:dataset_dw_index_max])
	
	return interped_data

f='Y:\ProcessFE\DailyWhiteBoard_2015_1Q ACTIVE.xlsx'
wb = load_workbook(filename = f,use_iterators=True, data_only=True)	
ws4 = wb.get_sheet_by_name(name="Runs")

porinclude=[334,376]#,417,419]#need to remove 417 and 418
porexculde=[299,303,305,307,316,331,332,339,344,361]
porrundata=[]
mr600rundata=[]
SPrundata=[]
otherrundata=[]
allrundata={}
rowcounter=0
for row in ws4.iter_rows():
	if rowcounter==0:
		labels=[row[0].value,row[6].value,row[3].value,row[10].value.encode('utf-8')]
	if row[10].value!=None and row[0].value!=None and rowcounter>0 and isinstance(row[0].value,float):
		added=False
		thedescr=row[10].value.encode('utf-8')#.decode('utf-8')
		#thedescr=thedescr.encode('CP1252')
		runnum=int(row[0].value)
		PCBEtool=row[6].value
		runlength=row[3].value
		allrundata[str(runnum)]=[PCBEtool,runlength]
		if runnum not in porexculde and re.search('POR',thedescr) or runnum in porinclude:
			if not re.search('POR-ish',thedescr):
				porrundata.append((runnum,PCBEtool,runlength,thedescr))
				allrundata[str(runnum)].append('POR')
				added=True
				#print 'POR run ', runnum
			else:
				otherrundata.append((runnum,PCBEtool,runlength,thedescr))
				allrundata[str(runnum)].append('other')
				added=True
			'''else:
				otherrundata.append((runnum,PCBEtool,runlength,thedescr))
				allrundata[str(runnum)].append('other')'''
		elif re.search('600',thedescr) and added==False:
			if not re.search('600m',thedescr):
				mr600rundata.append((runnum,PCBEtool,runlength,thedescr))
				allrundata[str(runnum)].append('MR600')
				added=True
		elif added==False:
			#del allrundata[str(runnum)]
			otherrundata.append((runnum,PCBEtool,runlength,thedescr))
			allrundata[str(runnum)].append('other')
		#if runnum==325:
		#	print thedescr.decode('utf-8')
	rowcounter+=1



runs = sorted(allrundata.keys())
#runsToRemove=['423','425','426']
#runs.remove('374')#374 XRF data is FUCKED UP
#runs.remove('351')#351 is 'POR-ish' pipecleaned fucking fucked up run
#runs.remove('405')#405 still sitting, only ran BE so far
'''for run in runs:
	if int(run)>419:
		runsToRemove.append(run)#latest runs right now need to be removed'''

'''for run in runsToRemove:
	runs.remove(run)'''

runsWithNoXRFfile = []
	
years = [13,14,15]
keys=['Cu','Ga','Mo','Se','Thickness','In','DW','Cu3','In3','Ga3']
##################get all the XRF data
XRFdata = {}
tempXRFdata = {}# to use to get raw data, interp to get full data
basepath = 'Y:\Experiment Summaries\Year 20'

printRunList = ['362','405','356']

print runs[98:]
for eachRun in runs[98:]:#runs:#runs[30:-2]
	eachRun = str(eachRun)
	noXRFfile = True
	print eachRun
	for year in years:
		runPath = basepath + str(year) + '\\' + 'S00' + str(eachRun) + '\\'
		if os.path.exists(runPath):
			#print runPath
			
			if eachRun in printRunList:
				print runPath + 'S*' + str(eachRun) + '*' + allrundata[eachRun][0] + '*' + 'xlsx'
			for f in glob.iglob(runPath + '*S*' + str(eachRun) + '*' + allrundata[eachRun][0] + '*' + 'xlsx'):
				print f
				if eachRun == '426':
					newFormat = True
				else:
					newFormat = False
				if noXRFfile:
					noXRFfile = False
					
					xrfFile = f

					wb = load_workbook(filename = xrfFile, use_iterators=True, data_only=True)	
					if wb.get_sheet_by_name(name = "MC01 XRF") != None:
						ws4 = wb.get_sheet_by_name(name = "MC01 XRF")
					elif wb.get_sheet_by_name(name = "MC02 XRF") != None:
						ws4 = wb.get_sheet_by_name(name = "MC02 XRF")
					else:
						wb.get_sheet_by_name(name = "XRF")
					'''
					row labels
					2: Cu
					3: Ga
					4: Mo
					5: Se
					6: Cu/III
					7: Thickness
					8: In
					9:  blank
					10: DW
					'''
					XRFdata[eachRun]={}
					tempXRFdata[eachRun]={}
					for key in keys:
						XRFdata[eachRun][key]=[]
						tempXRFdata[eachRun][key]=[]

						
					if newFormat: #426 and possible all newer runs have CuIII and Cu columns switched
						cuRow = 6
						gaRow = 3
						moRow = 4
						seRow = 5
						thRow = 7
						inRow = 8
						dwRow = 10
					else:
						cuRow = 2
						gaRow = 3
						moRow = 4
						seRow = 5
						thRow = 7
						inRow = 8
						dwRow = 10

					rowcounter=0
					for row in ws4.iter_rows():
						if rowcounter==0:
							pass#labels=[row[2].value,row[3].value,row[3].value,row[10].value.encode('utf-8')]
							rowcounter+=1
						if row[dwRow].value>=0 and row[dwRow].value<=allrundata[eachRun][1] and float(row[dwRow].value)!=-5.57:#-5.57 shows up in 389 when two rows are missing in 'web length' column
							if row[cuRow].value!=None:
								tempXRFdata[eachRun]['Cu'].append([row[cuRow].value,row[dwRow].value])
							if row[gaRow].value!=None:
								tempXRFdata[eachRun]['Ga'].append([row[gaRow].value,row[dwRow].value])
							if row[moRow].value!=None:
								tempXRFdata[eachRun]['Mo'].append([row[moRow].value,row[dwRow].value])
							if row[seRow].value!=None:
								tempXRFdata[eachRun]['Se'].append([row[seRow].value,row[dwRow].value])
							if row[thRow].value!=None:
								tempXRFdata[eachRun]['Thickness'].append([row[thRow].value,row[dwRow].value])
							if row[inRow].value!=None:
								tempXRFdata[eachRun]['In'].append([row[inRow].value,row[dwRow].value])
							XRFdata[eachRun]['DW'].append(row[dwRow].value)
			
	if not noXRFfile: #if found the XRF file
		for key in ['Cu','Ga','Mo','Se','Thickness','In']:
			tempXRFdata[eachRun][key]=np.array(tempXRFdata[eachRun][key],dtype='float64')
			XRFdata[eachRun][key]=np.interp(XRFdata[eachRun]['DW'],tempXRFdata[eachRun][key][:,1],tempXRFdata[eachRun][key][:,0])
		for count in range(len(XRFdata[eachRun]['DW'])):
			#print count
			XRFdata[eachRun]['Cu3'].append((XRFdata[eachRun]['Cu'][count])/(XRFdata[eachRun]['In'][count]+XRFdata[eachRun]['Ga'][count]))
			XRFdata[eachRun]['In3'].append((XRFdata[eachRun]['In'][count])/(XRFdata[eachRun]['Cu'][count]+XRFdata[eachRun]['Ga'][count]))
			XRFdata[eachRun]['Ga3'].append((XRFdata[eachRun]['Ga'][count])/(XRFdata[eachRun]['Cu'][count]+XRFdata[eachRun]['In'][count]))
	else:
		runsWithNoXRFfile.append(eachRun)
				
					

sorted_XRF_labels = ['Mo','Cu','In','Ga','Se','Cu3','In3','Ga3','Thickness'] #don't need DW cause it's in eff data


xrfruns = sorted(XRFdata.keys())



############import efficiency data

basepath='Y:\Nate'
efffile='eff data to 426.csv'


effCutoff=0
counter=0
alldata={}
tempeffdata=[]
dtype = [('DateTested', 'datetime64'), ('web_id', 'S10'), ('DW', 'float32'), 
('cell_area', 'float32'), ('substrateID', 'S10'), ('CW', 'float32'), 
('baked', 'int16'), ('mfgtype', 'S10'), ('rejectedreason', 'S10'),
('eff', 'float32'), ('voc', 'float32'), ('jsc', 'float32'), 
('ff', 'float32'), ('rs', 'float32'), ('rsh', 'float32')]
'''
todo: remove lines with blank or other for substrate id, DW, CW, eff, etc
daily check of eff data, append to full list of data with new
email out new stats

0=DateTested
1=web id
2=DW
3=cell area
4=substrate ID
5=CW
6=baked
7=mfgtype
8=rejectedreason
9=eff
10=voc
11=jsc
12=ff
13=rs
14=rsh'''
passed=0
notpassed=0
with open('Y:\\TASK FORCE - Performance drift\\list of POR runs\\raw eff data\\'+efffile, 'rb') as f:
    reader = csv.reader(f)
    for row in reader:
		if counter>0 and float(row[9])>=effCutoff:
			passed+=1
			tempeffdata.append(row)
		elif counter==0:
			labelrow=row
		else:
			tempeffdata.append(row)
			notpassed+=1
		counter+=1

#print 'passed: ', passed
#print 'not passed : ', notpassed
#print 'total yield : ', float(passed)/(float(notpassed+passed)), '%'

		
		
#tempeffdata=sorted(tempeffdata, key=lambda x: x[4])
#tempeffdata=np.array(tempeffdata, dtype=dtype)
#tempeffdata=np.sort(tempeffdata, order=['substrateID','DW'])

xrfRows = [0.46,0.5]

firstsubstrate=True
for row in tempeffdata:
	substrate=str(row[4]).lstrip('S00')
	if substrate in xrfruns and row[2]!='' and float(row[9])>=effCutoff:
		if float(row[5]) in xrfRows :
			if firstsubstrate:
				current_substrate=substrate
				alldata[current_substrate]=[]
				firstsubstrate=False
			else:
				if substrate!=current_substrate:
					current_substrate=substrate
					alldata[current_substrate]=[]
			alldata[current_substrate].append(row)

eff_data={}
for current_substrate in alldata.keys():
	alldata[current_substrate]=np.array(alldata[current_substrate])
	eff_data[current_substrate]={}
	eff_data[current_substrate]['DateTested']=alldata[current_substrate][:,0]
	eff_data[current_substrate]['Web ID']=alldata[current_substrate][:,1]
	eff_data[current_substrate]['DW']=np.array(alldata[current_substrate][:,2],dtype='float64')
	eff_data[current_substrate]['Cell Area']=np.array(alldata[current_substrate][:,3],dtype='float64')
	eff_data[current_substrate]['Substrate']=alldata[current_substrate][:,4]
	eff_data[current_substrate]['CW']=np.array(alldata[current_substrate][:,5],dtype='float64')
	eff_data[current_substrate]['Baked']=np.array(alldata[current_substrate][:,6],dtype='int8')
	eff_data[current_substrate]['MfgType']=alldata[current_substrate][:,7]
	eff_data[current_substrate]['RejectedReason']=alldata[current_substrate][:,8]
	eff_data[current_substrate]['Eff']=np.array(alldata[current_substrate][:,9],dtype='float64')
	eff_data[current_substrate]['Voc']=np.array(alldata[current_substrate][:,10],dtype='float64')
	eff_data[current_substrate]['Jsc']=np.array(alldata[current_substrate][:,11],dtype='float64')
	eff_data[current_substrate]['FF']=np.array(alldata[current_substrate][:,12],dtype='float64')
	eff_data[current_substrate]['Rs']=np.array(alldata[current_substrate][:,13],dtype='float64')
	eff_data[current_substrate]['Rsh']=np.array(alldata[current_substrate][:,14],dtype='float64')
substrate_list=sorted(alldata.keys())

print sorted(eff_data.keys())


interpdXRF={}
for substrate in eff_data.keys():
	if substrate in xrfruns:
		interpdXRF[substrate]={}
		for key in keys:
			#print substrate, key, len(XRFdata[substrate][key])
			interpdXRF[substrate][key]=interp_to_eff(eff_data[substrate]['DW'],XRFdata[substrate]['DW'],XRFdata[substrate][key])
	else:
		print substrate, ' not in xrf run list'

print 'no xrf data for ', [run for run in runsWithNoXRFfile]

with open('../xrf up to 426.csv', 'wb') as csvfile:
	spamwriter = csv.writer(csvfile, delimiter=',', quoting=csv.QUOTE_MINIMAL)
	print [key for key in sorted_XRF_labels]
	print [effkey for effkey in eff_data[xrfruns[0]].keys()]
	spamwriter.writerow([key for key in sorted_XRF_labels]+['run','BE/PC tool','run type']+[effkey for effkey in eff_data[xrfruns[0]].keys()])
	for substrate in sorted(eff_data.keys()):
		#print substrate, allrundata[substrate], [key for key in interpdXRF[substrate].keys()], [len(interpdXRF[substrate][key]) for key in interpdXRF[substrate].keys()]
		if substrate in xrfruns:
			for each in range(len(interpdXRF[substrate]['Cu'])):
				spamwriter.writerow([interpdXRF[substrate][key][each] for key in sorted_XRF_labels]+[substrate,allrundata[substrate][0],allrundata[substrate][2]]+[eff_data[substrate][effkey][each] for effkey in eff_data[substrate].keys()])

